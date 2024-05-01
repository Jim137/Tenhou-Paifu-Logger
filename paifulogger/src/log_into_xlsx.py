import re
from typing import cast

import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet

from .i18n import LocalStr
from .Paifu import Paifu


def log_into_xlsx(paifu: Paifu, local_lang: LocalStr, output: str):
    """
    Log the paifu data into xlsx file.

    Parameters
    ----------
    paifu: Paifu
        The paifu data.
    local_lang: LocalStr
        The localized string.
    output: str
        The output directory.
    """

    try:
        if paifu.player_num == 3:
            paifu_str = local_lang.sanma + local_lang.paifu
        else:
            paifu_str = local_lang.yonma + local_lang.paifu
        path = f"{output}/{local_lang.paifu}/{paifu_str}.xlsx"
        wb = xl.load_workbook(path)

        # Prevent MyPy warning
        sheet = cast(Worksheet, wb.active)
        # Remove final statistics
        sheet.delete_rows(sheet.max_row)
    except FileNotFoundError:
        # Create a new workbook
        wb = xl.Workbook()
        sheet = cast(Worksheet, wb.active)
        sheet.append(
            [
                local_lang.date,
                local_lang.plc,
                local_lang.paifu,
                local_lang.remark,
                local_lang.preR,
                local_lang.r_change,
                local_lang.round_num,
                local_lang.win,
                local_lang.deal_in,
            ]
        )
        # set column width
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["C"].width = 71
        sheet.column_dimensions["E"].width = local_lang.excelE
    # Append new paifu data to the sheet
    sheet.append(
        [
            paifu.time,
            paifu.get_place(paifu.ban),
            paifu.url,
            "",
            float(paifu.r[paifu.ban]),
            paifu.rate_change,
            paifu.get_round_num(),
            paifu.get_win_num(paifu.ban),
            paifu.get_deal_in_num(paifu.ban),
        ]
    )
    sheet["C" + str(sheet.max_row)].style = "Hyperlink"
    # Add final statistics
    sheet.append(
        [
            local_lang.avg_plc,
            f"=average(B2:B{sheet.max_row})",
            "",
            local_lang.win_rate,
            f"=sum(H2:H{sheet.max_row})/sum(G2:G{sheet.max_row})",
            "",
            local_lang.deal_in_rate,
            f"=sum(I2:I{sheet.max_row})/sum(G2:G{sheet.max_row})",
        ]
    )
    sheet["E" + str(sheet.max_row)].style = "Percent"
    sheet["H" + str(sheet.max_row)].style = "Percent"
    wb.save(path)
    print(
        "xlsx: "
        + local_lang.hint_record1
        + re.findall(r"\d{10}gm-\w{4}-\w{4}-\w{8}&tw=\d", paifu.url)[0]
        + local_lang.hint_record2
    )
    return None
