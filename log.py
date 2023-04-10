import openpyxl as xl
from datetime import datetime
import os
import urllib.request
import sys
from src import *


if __name__ == '__main__':
    if len(sys.argv) > 1:
        lang = sys.argv[1]
    else:
        lang = 'en'
    local_str = localized_str(lang)

    if not os.path.isdir(f'./{local_str.haifu}/'):
        os.makedirs(f'./{local_str.haifu}/')

    url = input(local_str.hint_input)
    try:
        haifu = get_haifu(url, local_str)
        p = get_place(haifu, haifu.ban)
        r = haifu.r[haifu.ban]
        try:
            wb = xl.load_workbook(f'{local_str.haifu}.xlsx')
            sheet = wb.active
            sheet.delete_rows(sheet.max_row)
        except FileNotFoundError:
            wb = xl.Workbook()
            sheet = wb.active
            sheet.append([local_str.date, local_str.plc,
                         local_str.haifu, local_str.remark, local_str.preR])
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['C'].width = 71
            sheet.column_dimensions['E'].width = local_str.excelE
        sheet.append(
            [datetime.strptime(url[26:36], "%Y%m%d%H"), p, url, '', float(r)])
        sheet['C'+str(sheet.max_row)].style = "Hyperlink"
        sheet.append(
            [local_str.avg_plc, '=average(B2:B'+str(sheet.max_row)+')'])
        wb.save(f'{local_str.haifu}.xlsx')
        print(local_str.hint_record1+url[26:]+local_str.hint_record2)
    except urllib.error.URLError:
        print(local_str.hint_url)
    except ValueError:
        print(local_str.hint_tw)
