import openpyxl as xl
from datetime import datetime
from .get_place import get_place
from .Paifu import Paifu
from .local import local_str


def log_into_xlsx(paifu: Paifu, local_str: local_str):
    try:
        if paifu.player_num == 3:
            paifu_str = local_str.sanma + local_str.paifu
        else:
            paifu_str = local_str.yonma + local_str.paifu
        wb = xl.load_workbook(f'./{local_str.paifu}/{paifu_str}.xlsx')
        sheet = wb.active
        sheet.delete_rows(sheet.max_row)
    except FileNotFoundError:
        wb = xl.Workbook()
        sheet = wb.active
        sheet.append([local_str.date, local_str.plc,
                      local_str.paifu, local_str.remark, local_str.preR])
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['C'].width = 71
        sheet.column_dimensions['E'].width = local_str.excelE
    sheet.append(
        [datetime.strptime(paifu.url[26:36], "%Y%m%d%H"), get_place(paifu, paifu.ban), paifu.url, '', float(paifu.r[paifu.ban])])
    sheet['C'+str(sheet.max_row)].style = "Hyperlink"
    sheet.append(
        [local_str.avg_plc, '=average(B2:B'+str(sheet.max_row)+')'])
    wb.save(f'./{local_str.paifu}/{paifu_str}.xlsx')
    print('xlsx: '+local_str.hint_record1+paifu.url[26:]+local_str.hint_record2)
    return None
